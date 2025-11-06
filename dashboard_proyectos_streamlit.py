# =============================================
# 📊 DASHBOARD DE PROYECTOS - STREAMLIT (FUNDACIÓN CUENCAS SAGRADAS)
# =============================================
import warnings
warnings.filterwarnings("ignore")

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from sqlalchemy import create_engine, text
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# =========================
# CONFIGURACIÓN INICIAL
# =========================
st.set_page_config(page_title="FUNDACIÓN CUENCAS SAGRADAS - Dashboard", layout="wide")
st.title("FUNDACIÓN CUENCAS SAGRADAS")
st.subheader("Sistema Indígena de Monitoreo, Seguimiento, Evaluación y Aprendizaje")
st.markdown("Panel interactivo: ejecución presupuestaria y física — filtra y descarga tus datos.")

# =========================
# CONEXIÓN A BASE DE DATOS
# =========================
DB_PATH = "proyectos_backup.db"
ENGINE = create_engine(f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False})

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS proyectos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT,
    nombre_proyecto TEXT,
    pais TEXT,
    provincia TEXT,
    canton TEXT,
    pueblo TEXT,
    latitud REAL,
    longitud REAL,
    objetivo_general TEXT,
    beneficiarios_hombres INTEGER,
    beneficiarios_mujeres INTEGER,
    beneficiarios_glbti INTEGER,
    total_beneficiarios INTEGER,
    plazo_ejecucion TEXT,
    fecha_inicio TEXT,
    fecha_cierre TEXT,
    monto_financiamiento REAL,
    nivel_objetivo TEXT,
    indicadores TEXT,
    linea_base TEXT,
    meta_final REAL,
    meta_alcanzada REAL,
    presupuesto_planificado REAL,
    presupuesto_devengado REAL,
    frecuencia TEXT,
    medios_verificacion TEXT
)
"""
with ENGINE.connect() as conn:
    conn.execute(text(CREATE_SQL))

# =========================
# FUNCIONES BASE DE DATOS
# =========================
@st.cache_data(ttl=30)
def cargar_datos():
    try:
        df = pd.read_sql_query("SELECT * FROM proyectos", ENGINE)
    except Exception:
        cols = [
            "id","codigo","nombre_proyecto","pais","provincia","canton","pueblo",
            "latitud","longitud","objetivo_general","beneficiarios_hombres",
            "beneficiarios_mujeres","beneficiarios_glbti","total_beneficiarios",
            "plazo_ejecucion","fecha_inicio","fecha_cierre","monto_financiamiento",
            "nivel_objetivo","indicadores","linea_base","meta_final","meta_alcanzada",
            "presupuesto_planificado","presupuesto_devengado","frecuencia","medios_verificacion"
        ]
        df = pd.DataFrame(columns=cols)
    # parsear fechas si están presentes
    if "fecha_inicio" in df.columns:
        df["fecha_inicio"] = pd.to_datetime(df["fecha_inicio"], errors="coerce")
    if "fecha_cierre" in df.columns:
        df["fecha_cierre"] = pd.to_datetime(df["fecha_cierre"], errors="coerce")
    return df

# función que crea un Excel (BytesIO) desde un DataFrame
def crear_excel_en_memoria(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Proyectos"

    # volcar DF
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # formato encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ajustar anchos sin errores por MergedCell
    for col in ws.columns:
        try:
            first_cell = col[0]
            col_letter = get_column_letter(first_cell.column)
        except Exception:
            continue
        max_length = 0
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# =========================
# CARGA Y FILTROS
# =========================
df = cargar_datos()

if df.empty:
    st.warning("⚠️ No hay proyectos en la base de datos. Usa el formulario para agregar registros.")
    st.stop()

# Sidebar - filtros relevantes
st.sidebar.header("🔎 Filtros")
paises = ["Todos"] + sorted(df["pais"].dropna().unique().tolist())
provincias = ["Todos"] + sorted(df["provincia"].dropna().unique().tolist())
pueblos = ["Todos"] + sorted(df["pueblo"].dropna().unique().tolist())
niveles = ["Todos"] + sorted(df["nivel_objetivo"].dropna().unique().tolist())

pais_sel = st.sidebar.selectbox("País", paises)
prov_sel = st.sidebar.selectbox("Provincia", provincias)

# Pueblo: multiselect con opción "Todos"
pueblo_sel = st.sidebar.multiselect("Pueblo (selecciona uno o varios)", pueblos, default=["Todos"] if "Todos" in pueblos else pueblos)

nivel_sel = st.sidebar.multiselect("Nivel objetivo (múltiple)", niveles, default=["Todos"] if "Todos" in niveles else niveles)

rango_fecha = st.sidebar.date_input(
    "Rango fecha inicio",
    value=(
        df["fecha_inicio"].min().date() if pd.notna(df["fecha_inicio"].min()) else datetime.today().date(),
        df["fecha_inicio"].max().date() if pd.notna(df["fecha_inicio"].max()) else datetime.today().date()
    )
)

monto_min = float(df["monto_financiamiento"].fillna(0).min()) if "monto_financiamiento" in df.columns else 0.0
monto_max = float(df["monto_financiamiento"].fillna(0).max()) if "monto_financiamiento" in df.columns else 0.0
monto_sel = st.sidebar.slider("Monto financiación", min_value=0.0, max_value=max(1.0, monto_max), value=(monto_min, monto_max))

# aplicar filtros
df_filtrado = df.copy()
if pais_sel != "Todos":
    df_filtrado = df_filtrado[df_filtrado["pais"] == pais_sel]
if prov_sel != "Todos":
    df_filtrado = df_filtrado[df_filtrado["provincia"] == prov_sel]

# aplicar filtro pueblo (multiselección), respetando la opción "Todos"
if pueblo_sel and "Todos" not in pueblo_sel:
    df_filtrado = df_filtrado[df_filtrado["pueblo"].isin(pueblo_sel)]

# aplicar filtro nivel objetivo (multiselección)
if nivel_sel and "Todos" not in nivel_sel:
    df_filtrado = df_filtrado[df_filtrado["nivel_objetivo"].isin(nivel_sel)]

# aplicar rango de fechas
try:
    start_date, end_date = rango_fecha
    df_filtrado = df_filtrado[(df_filtrado["fecha_inicio"] >= pd.to_datetime(start_date)) & (df_filtrado["fecha_inicio"] <= pd.to_datetime(end_date))]
except Exception:
    pass

# aplicar filtro monto
if "monto_financiamiento" in df_filtrado.columns:
    df_filtrado = df_filtrado[(df_filtrado["monto_financiamiento"].fillna(0) >= monto_sel[0]) & (df_filtrado["monto_financiamiento"].fillna(0) <= monto_sel[1])]

# asegurar columnas numéricas que vamos a usar
for col in ["presupuesto_planificado", "presupuesto_devengado", "meta_final", "meta_alcanzada", "total_beneficiarios", "monto_financiamiento"]:
    if col in df_filtrado.columns:
        df_filtrado[col] = pd.to_numeric(df_filtrado[col], errors="coerce").fillna(0)

# =========================
# INDICADORES CLAVE (KPI)
# =========================
st.header("📈 Indicadores Clave")
col1, col2, col3, col4 = st.columns(4)
col1.metric("📁 Proyectos (filtrados)", len(df_filtrado))
col2.metric("👥 Beneficiarios (total)", int(df_filtrado["total_beneficiarios"].sum() if "total_beneficiarios" in df_filtrado.columns else 0))
col3.metric("💰 Presupuesto Planificado (total)", f"${df_filtrado['presupuesto_planificado'].sum():,.2f}" if "presupuesto_planificado" in df_filtrado.columns else "$0.00")
col4.metric("💸 Presupuesto Devengado (total)", f"${df_filtrado['presupuesto_devengado'].sum():,.2f}" if "presupuesto_devengado" in df_filtrado.columns else "$0.00")

# =========================
# EJECUCIÓN PRESUPUESTARIA GLOBAL
# =========================
st.subheader("📊 Ejecución Presupuestaria Global")

total_plan = df_filtrado["presupuesto_planificado"].sum() if "presupuesto_planificado" in df_filtrado.columns else 0.0
total_dev = df_filtrado["presupuesto_devengado"].sum() if "presupuesto_devengado" in df_filtrado.columns else 0.0
ejec_pres = (total_dev / total_plan * 100) if total_plan > 0 else 0.0

c1, c2 = st.columns(2)
c1.metric("Presupuesto Planificado (total)", f"${total_plan:,.2f}")
c2.metric("Presupuesto Devengado (total)", f"${total_dev:,.2f}")

# Gauge y explicación
g1, g2 = st.columns([1, 2])
fig_gauge_pres = go.Figure(go.Indicator(
    mode="gauge+number",
    value=ejec_pres,
    title={"text": "Ejecución Presupuestaria (%)"},
    gauge={
        "axis": {"range": [0, 120]},
        "steps": [
            {"range": [0, 50], "color": "lightcoral"},
            {"range": [50, 80], "color": "khaki"},
            {"range": [80, 120], "color": "lightgreen"}
        ],
        "bar": {"color": "darkgreen"}
    }
))
g1.plotly_chart(fig_gauge_pres, use_container_width=True)
g2.write(f"**Cálculo:** (Presupuesto devengado / Presupuesto planificado) × 100 = **{ejec_pres:.2f}%**")

# =========================
# EJECUCIÓN FÍSICA GLOBAL
# =========================
st.subheader("🎯 Ejecución Física Global")
total_meta = df_filtrado["meta_final"].sum() if "meta_final" in df_filtrado.columns else 0.0
total_alc = df_filtrado["meta_alcanzada"].sum() if "meta_alcanzada" in df_filtrado.columns else 0.0
ejec_fis = (total_alc / total_meta * 100) if total_meta > 0 else 0.0

st.write(f"**Ejecución Física (global):** {ejec_fis:.2f}%")
fig_gauge_fis = go.Figure(go.Indicator(
    mode="gauge+number",
    value=ejec_fis,
    title={"text": "Ejecución Física (%)"},
    gauge={
        "axis": {"range": [0, 120]},
        "steps": [
            {"range": [0, 50], "color": "lightcoral"},
            {"range": [50, 80], "color": "khaki"},
            {"range": [80, 120], "color": "lightgreen"}
        ],
        "bar": {"color": "darkblue"}
    }
))
st.plotly_chart(fig_gauge_fis, use_container_width=True)

# =========================
# GRÁFICOS COMPARATIVOS
# =========================
st.markdown("---")
st.subheader("📊 Comparativos")

# comparativo presupuestario por proyecto (mostrar % por proyecto)
detalle = df_filtrado.copy()
detalle["Ejecución Presupuestaria (%)"] = detalle.apply(
    lambda r: (r["presupuesto_devengado"] / r["presupuesto_planificado"] * 100) if (r.get("presupuesto_planificado", 0) > 0) else 0.0,
    axis=1
)
detalle["Ejecución Física (%)"] = detalle.apply(
    lambda r: (r["meta_alcanzada"] / r["meta_final"] * 100) if (r.get("meta_final", 0) > 0) else 0.0,
    axis=1
)

# barra: ejec pres vs ejec fis (top 30 por pres)
if not detalle.empty:
    detalle_sorted = detalle.sort_values("Ejecución Presupuestaria (%)", ascending=False).head(30)
    fig_comp = px.bar(
        detalle_sorted,
        x="nombre_proyecto",
        y=["Ejecución Presupuestaria (%)", "Ejecución Física (%)"],
        barmode="group",
        title="Ejecución Presupuestaria y Física por Proyecto (Top 30 por Presupuesto)",
        labels={"value": "Porcentaje (%)", "nombre_proyecto": "Proyecto"}
    )
    fig_comp.update_layout(xaxis_tickangle=-45, height=480, margin=dict(t=60))
    st.plotly_chart(fig_comp, use_container_width=True)
else:
    st.info("No hay datos para graficar comparativo.")

# Agregado por provincia (ejecución presupuestaria media)
st.subheader("🗺️ Promedio de Ejecución por Provincia")
if "provincia" in detalle.columns and not detalle["provincia"].dropna().empty:
    prov_group = detalle.groupby("provincia").agg({
        "Ejecución Presupuestaria (%)": "mean",
        "Ejecución Física (%)": "mean",
        "presupuesto_planificado": "sum",
        "presupuesto_devengado": "sum",
        "id": "count"
    }).rename(columns={"id": "num_proyectos"}).reset_index()
    prov_group = prov_group.sort_values("Ejecución Presupuestaria (%)", ascending=False).head(30)
    fig_prov = px.bar(
        prov_group,
        x="provincia",
        y=["Ejecución Presupuestaria (%)", "Ejecución Física (%)"],
        barmode="group",
        title="Promedio de Ejecución por Provincia (Top 30)",
        labels={"value": "Porcentaje (%)", "provincia": "Provincia"}
    )
    fig_prov.update_layout(xaxis_tickangle=-45, height=420)
    st.plotly_chart(fig_prov, use_container_width=True)
else:
    st.info("No hay datos por provincia para el agregado.")

# =========================
# MAPA DE PROYECTOS
# =========================
st.markdown("---")
st.subheader("🗺️ Mapa de Proyectos")
if {"latitud", "longitud"}.issubset(df_filtrado.columns) and not df_filtrado[["latitud", "longitud"]].dropna().empty:
    map_df = df_filtrado[["latitud", "longitud", "nombre_proyecto", "codigo"]].dropna().rename(columns={"latitud": "lat", "longitud": "lon"})
    st.map(map_df.loc[:, ["lat", "lon"]])
else:
    st.info("No hay coordenadas válidas para mostrar en el mapa.")

# =========================
# DETALLE FINAL + EXPORTACIÓN (AL FINAL DEL ARCHIVO)
# =========================
st.markdown("---")
st.subheader("📋 Detalle final de Proyectos (filtrado)")

# columnas a mostrar (ordenadas) - ahora incluimos meta_final y meta_alcanzada en el listado final como pediste
cols_to_show = [
    "id", "codigo", "nombre_proyecto", "pais", "provincia", "canton", "pueblo",
    "presupuesto_planificado", "presupuesto_devengado", "Ejecución Presupuestaria (%)",
    "meta_final", "meta_alcanzada", "Ejecución Física (%)", "monto_financiamiento", "total_beneficiarios"
]
cols_show_existing = [c for c in cols_to_show if c in detalle.columns]

if cols_show_existing:
    df_detalle_final = detalle[cols_show_existing].sort_values("Ejecución Presupuestaria (%)", ascending=False).reset_index(drop=True)
    st.dataframe(df_detalle_final, use_container_width=True)
else:
    st.info("No hay columnas para mostrar en el detalle final.")

# Botón exportar CSV
csv = df_detalle_final.to_csv(index=False).encode("utf-8")
st.download_button("⬇️ Exportar CSV (filtrado)", csv, "proyectos_filtrados.csv", "text/csv")

# Botón exportar Excel (genera buffer al hacer clic)
if st.button("📤 Generar y Descargar Excel (filtrado)"):
    excel_buf = crear_excel_en_memoria(df_detalle_final)
    st.download_button(
        label="⬇️ Descargar Excel",
        data=excel_buf.getvalue(),
        file_name=f"reporte_proyectos_cuencas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# FOOTER
# =========================
st.markdown("---")
st.caption("FUNDACIÓN CUENCAS SAGRADAS | Ejecuta con: streamlit run dashboard_proyectos_cuencas.py")